#!/usr/bin/env python3
import openpyxl
FIN = '/Users/russellteter/Library/Application Support/Claude/local-agent-mode-sessions/fa5c2f7e-5fb9-4e29-a76c-e706355df1a1/f2ae62ca-b383-441b-9a66-f02d2b790532/local_54f4fee2-d572-4c3a-8af3-bdd6e183d27d/uploads/Class Cash Forecast as of May 10 2026 (1).xlsx'

wb = openpyxl.load_workbook(FIN, data_only=True)
ws = wb['Class Cash Flows']

# Print column headers (week ending dates in row 6)
weeks = []
for c in range(3, ws.max_column+1):
    v = ws.cell(6,c).value
    if v: weeks.append((c, str(v)[:10]))
print(f"Weeks in forecast: {len(weeks)}")
print(f"First week: col {weeks[0][0]} = {weeks[0][1]}")
print(f"Last week:  col {weeks[-1][0]} = {weeks[-1][1]}")

# Find actuals vs forecast cutoff (row 5)
row5 = [(c, ws.cell(5,c).value) for c in range(3, ws.max_column+1) if ws.cell(5,c).value]
last_actual = None; first_forecast = None
for c, v in row5:
    if 'Actual' in str(v): last_actual = c
    elif 'Forecast' in str(v) and first_forecast is None: first_forecast = c
print(f"Actuals end at col {last_actual} ({weeks[last_actual-3][1] if last_actual else None})")
print(f"Forecast starts at col {first_forecast} ({weeks[first_forecast-3][1] if first_forecast else None})")

# Dump key rows fully  
print("\n--- Row B labels with non-zero values ---")
for r in range(7, 45):
    label = ws.cell(r, 2).value
    if not label: continue
    vals = [ws.cell(r,c).value for c in range(3, ws.max_column+1)]
    total = sum(v for v in vals if isinstance(v,(int,float)))
    nonzero = sum(1 for v in vals if isinstance(v,(int,float)) and v != 0)
    if nonzero > 0:
        print(f"R{r:>2} | {str(label)[:38]:<38} | sum={total:>16,.0f} | nz_weeks={nonzero}")

# Specifically dump AWS Fees row across all weeks
print("\n--- AWS Fees (row 22) per week ---")
for c, wk in weeks:
    v = ws.cell(22, c).value
    typ = ws.cell(5, c).value or ''
    if isinstance(v,(int,float)) and v != 0:
        print(f"  {wk}  [{typ:<8}]  ${v:>14,.0f}")

# AWS Fees forecast subset (Jun-Aug 2026)
print("\n--- AWS Fees Finance forecast 2026 ---")
import datetime
total_2026 = 0
total_jun_aug = 0
fc_only = 0
for c, wk in weeks:
    v = ws.cell(22, c).value
    typ = str(ws.cell(5, c).value or '')
    if isinstance(v,(int,float)):
        total_2026 += v
        if 'Forecast' in typ: fc_only += v
        if wk >= '2026-06-01' and wk <= '2026-08-31':
            total_jun_aug += v
print(f"Total 2026 AWS Fees: ${total_2026:,.0f}")
print(f"Forecast-only AWS Fees: ${fc_only:,.0f}")
print(f"Jun-Aug 2026 AWS Fees: ${total_jun_aug:,.0f}")

# CoSo Hosting
print("\n--- CoSo Hosting (row 24) Finance forecast 2026 ---")
total = 0
for c, wk in weeks:
    v = ws.cell(24, c).value
    if isinstance(v,(int,float)):
        total += v
        if v != 0:
            typ = str(ws.cell(5, c).value or '')
            print(f"  {wk}  [{typ:<8}]  ${v:>14,.0f}")
print(f"Total 2026 CoSo Hosting: ${total:,.0f}")
