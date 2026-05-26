#!/usr/bin/env python3
"""Pull AR rows + the trough week math from Finance file."""
import openpyxl
FIN = '/Users/russellteter/Library/Application Support/Claude/local-agent-mode-sessions/fa5c2f7e-5fb9-4e29-a76c-e706355df1a1/f2ae62ca-b383-441b-9a66-f02d2b790532/local_54f4fee2-d572-4c3a-8af3-bdd6e183d27d/uploads/Class Cash Forecast as of May 10 2026 (1).xlsx'

wb = openpyxl.load_workbook(FIN, data_only=True)
ws = wb['Class Cash Flows']

# Build week index
weeks = [(c, str(ws.cell(6,c).value)[:10], str(ws.cell(5,c).value or '')) for c in range(3, ws.max_column+1)]

# Print AR rows + ending cash row across W26-W34 (around the trough)
trough_window = [w for w in weeks if w[1] >= '2026-06-15' and w[1] <= '2026-08-15']
print("Trough window:")
print(f"  {'Week_End':<12} {'Type':<10}")
for c, wk, typ in trough_window:
    print(f"  {wk:<12} {typ:<10} col={c}")

# Find ending cash row
print("\n--- Finding 'Ending Cash' row ---")
for r in range(40, 46):
    label = ws.cell(r, 2).value
    print(f"R{r}: {label}")

# Try to identify the ending cash balance row
# Looking at row 42 = Beginning Cash; row 43 = Net Cash Flow; row 44 = Ending Cash probably
print()
ending_row = None
for r in range(40, 46):
    label = str(ws.cell(r, 2).value or '').lower()
    if 'ending' in label or 'eom' in label:
        ending_row = r
        print(f"Ending cash candidate row {r}: {ws.cell(r,2).value}")

# Force-look at row 44 and 45
for r in [42, 43, 44, 45]:
    label = ws.cell(r, 2).value
    if label:
        print(f"R{r} label: {label}")
        vals_around_trough = []
        for c, wk, typ in trough_window:
            v = ws.cell(r, c).value
            if isinstance(v,(int,float)):
                vals_around_trough.append((wk, v))
        for wk, v in vals_around_trough:
            print(f"  {wk}  ${v:>14,.0f}")

# Now pull AR rows (10-14) across trough window
print("\n--- AR rows W25-W33 (around trough) ---")
for r in [10, 11, 12, 13, 14]:
    label = ws.cell(r, 2).value
    if not label: continue
    print(f"\n{label}")
    for c, wk, typ in trough_window:
        v = ws.cell(r, c).value
        if isinstance(v,(int,float)) and v != 0:
            print(f"  {wk}  [{typ:<8}]  ${v:>14,.0f}")
