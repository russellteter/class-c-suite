#!/usr/bin/env python3
import openpyxl
FIN = '/Users/russellteter/Library/Application Support/Claude/local-agent-mode-sessions/fa5c2f7e-5fb9-4e29-a76c-e706355df1a1/f2ae62ca-b383-441b-9a66-f02d2b790532/local_54f4fee2-d572-4c3a-8af3-bdd6e183d27d/uploads/Class Cash Forecast as of May 10 2026 (1).xlsx'

wb = openpyxl.load_workbook(FIN, data_only=True)
ws = wb['Class Cash Flows']
print(f"Dim: {ws.max_row} rows x {ws.max_column} cols")

# Print the first few rows to find headers
print("\n--- First 8 rows (headers) ---")
for r in range(1, 9):
    row = []
    for c in range(1, min(ws.max_column, 30)+1):
        v = ws.cell(r,c).value
        if v is None: row.append('')
        elif isinstance(v, float): row.append(f"{v:,.0f}")
        else: row.append(str(v)[:18])
    if any(row):
        print(f"R{r:>2}: " + " | ".join(f"{x[:15]:<15}" for x in row[:20]))

# Now find rows with 'host', 'aws', 'collab', 'collect', 'AR' in col B or A
print("\n--- Rows mentioning hosting / AWS / collections / AR ---")
keywords = ['host','aws','collab','collect','AR ','recover','BME','blackboard','knox','TSA']
for r in range(1, ws.max_row+1):
    cell_a = str(ws.cell(r,1).value or '').lower()
    cell_b = str(ws.cell(r,2).value or '').lower()
    if any(k.lower() in cell_a or k.lower() in cell_b for k in keywords):
        # Print this row's first 12 columns
        row = []
        for c in range(1, 13):
            v = ws.cell(r,c).value
            if v is None: row.append('')
            elif isinstance(v, float): row.append(f"{v:>12,.0f}")
            else: row.append(str(v)[:20])
        print(f"R{r:>3}: " + " | ".join(row))
