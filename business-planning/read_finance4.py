#!/usr/bin/env python3
"""Pull ALL row values for the trough-week and W29 (the AWS payment week)."""
import openpyxl
FIN = '/Users/russellteter/Library/Application Support/Claude/local-agent-mode-sessions/fa5c2f7e-5fb9-4e29-a76c-e706355df1a1/f2ae62ca-b383-441b-9a66-f02d2b790532/local_54f4fee2-d572-4c3a-8af3-bdd6e183d27d/uploads/Class Cash Forecast as of May 10 2026 (1).xlsx'

wb = openpyxl.load_workbook(FIN, data_only=True)
ws = wb['Class Cash Flows']

# Critical weeks for the trough
TROUGH_WEEKS = ['2026-07-12','2026-07-19','2026-07-26','2026-08-02']  
cols_to_show = []
for c in range(3, ws.max_column+1):
    wk = str(ws.cell(6,c).value)[:10]
    if wk in TROUGH_WEEKS:
        cols_to_show.append((c, wk))

print(f"Cols: {cols_to_show}")
print(f"\n{'Row':<5}{'Label':<42}" + "".join(f"{wk:>14}" for c,wk in cols_to_show))
print("-"*(47 + 14*len(cols_to_show)))
for r in range(8, 46):
    label = ws.cell(r,2).value
    if not label: continue
    label = str(label).strip()
    vals = []
    any_val = False
    for c, wk in cols_to_show:
        v = ws.cell(r, c).value
        if isinstance(v,(int,float)):
            if v != 0: any_val = True
            vals.append(f"{v:>14,.0f}")
        else:
            vals.append(f"{'':>14}")
    # Print rows with any non-zero in trough window OR rollup rows
    if any_val or 'Total' in label or 'Ending' in label or 'Beginning' in label or 'Net Cash' in label:
        print(f"{r:<5}{label[:41]:<42}" + "".join(vals))
