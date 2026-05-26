#!/usr/bin/env python3
"""Read NS_Variance and Cost_Levers from cash model + Finance cash forecast."""
import openpyxl

MODEL = '/Users/russellteter/Documents/Claude/Projects/Business Planning/Class_Cash_Lever_Model_v5_2026-05-18.xlsx'
FIN   = '/Users/russellteter/Library/Application Support/Claude/local-agent-mode-sessions/fa5c2f7e-5fb9-4e29-a76c-e706355df1a1/f2ae62ca-b383-441b-9a66-f02d2b790532/local_54f4fee2-d572-4c3a-8af3-bdd6e183d27d/uploads/Class Cash Forecast as of May 10 2026 (1).xlsx'

def dump_sheet(path, sheet_name, max_rows=80, max_cols=16):
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"\n=== {path.split('/')[-1]}  [sheet: {sheet_name}] ===")
    if sheet_name not in wb.sheetnames:
        print(f"  Not found. Available: {wb.sheetnames}")
        return
    ws = wb[sheet_name]
    print(f"  Dimensions: {ws.max_row} rows x {ws.max_column} cols")
    for r in range(1, min(ws.max_row, max_rows)+1):
        row = []
        for c in range(1, min(ws.max_column, max_cols)+1):
            v = ws.cell(r,c).value
            if v is None: row.append('')
            elif isinstance(v, float): row.append(f"{v:,.2f}")
            else: row.append(str(v)[:30])
        if any(row):
            print(f"  R{r:>2}: " + " | ".join(f"{x[:18]:<18}" for x in row))

def list_sheets(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"\n{path.split('/')[-1]} sheets: {wb.sheetnames}")

list_sheets(MODEL)
list_sheets(FIN)

for s in ['05_AWS_Analysis','06_NS_Variance','03_Cost_Levers']:
    dump_sheet(MODEL, s, max_rows=70, max_cols=16)
